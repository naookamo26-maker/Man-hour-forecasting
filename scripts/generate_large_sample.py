"""実データ規模のサンプルデータを生成する。

想定している実データの姿
    完了案件      30 件
    進行中案件     5 件(実績が途中まで。終了日は計画値でずれうる)
    予測対象       1 件(実績なし)
    期間          12〜36ヶ月
    工数          200〜3000 人月
    行程          20 個 / 行程グループ 10 個
    マイルストーン  全案件に α版・β版(完了案件にはマスターアップも)

出力先は --out で指定する(既定 data_large/)。
data/ の小さいサンプルは高速な動作確認用に残す。

生成の考え方は scripts/generate_sample_data.py と同じ。
正準時間軸上に行程グループごとの密度を定義し、α版・β版の直前に山を置き、
案件ごとにマイルストーン位置をばらつかせてワープで実時間軸に落とす。
位置合わせが効くかどうかを leave-one-out で測れる構造にしてある。

進行中案件については、実データで起きることを意図的に埋め込んでいる。
  - 実績が途中で切れる(α版通過直後 〜 マスターアップ直前まで様々)
  - 終了日は計画値であり、実際の完了月とはずれる
  - まだ来ていないマイルストーンは計画日として入っている
"""

from __future__ import annotations

import argparse
import datetime as dt
import os

import numpy as np
import pandas as pd

from generate_sample_data import (HEADER_FILL, HEADER_FONT, beta_pdf,
                                  month_edges_t, month_range, write_sheet)
from openpyxl import Workbook

HOURS_PER_MM = 160

# --------------------------------------------------------------------------
# 行程グループ 10 個 / 行程 20 個
# --------------------------------------------------------------------------
PHASES = {
    "企画・プリプロ":     [("企画立案", 0.45), ("プロトタイプ制作", 0.55)],
    "ゲームデザイン":     [("レベルデザイン", 0.55), ("バトルデザイン", 0.45)],
    "シナリオ":           [("シナリオ執筆", 0.60), ("テキスト実装", 0.40)],
    "プログラム":         [("クライアント実装", 0.50), ("サーバ実装", 0.25),
                           ("エンジン開発", 0.25)],
    "アート3D":           [("キャラクターモデリング", 0.55), ("背景モデリング", 0.45)],
    "アート2D":           [("2Dイラスト", 0.55), ("UIアート", 0.45)],
    "アニメーション・演出": [("アニメーション", 0.60), ("エフェクト", 0.40)],
    "サウンド":           [("BGM制作", 0.60), ("SE制作", 0.40)],
    "QA・デバッグ":       [("デバッグ", 0.70), ("QA計画・進行", 0.30)],
    "PM・運営":           [("プロジェクト管理", 1.00)],
}
GROUPS = list(PHASES)

MAJOR = {
    "企画・プリプロ": "上流", "ゲームデザイン": "上流", "シナリオ": "上流",
    "プログラム": "制作", "アート3D": "制作", "アート2D": "制作",
    "アニメーション・演出": "制作", "サウンド": "制作",
    "QA・デバッグ": "検証", "PM・運営": "管理",
}

# 正準時間軸上の形状: (beta_a, beta_b, [(山の中心, 高さ, 幅), ...])
# α版(0.50)・β版(0.76)の直前に山を置く。位置合わせの効果が測れるようにするため。
SHAPE = {
    "企画・プリプロ":       (1.4, 6.0, []),
    "ゲームデザイン":       (2.0, 3.4, [(0.50, 0.55, 0.05)]),
    "シナリオ":             (1.9, 3.8, [(0.50, 0.40, 0.05)]),
    "プログラム":           (2.6, 2.4, [(0.50, 0.80, 0.045), (0.76, 0.45, 0.04)]),
    "アート3D":             (3.2, 2.6, [(0.76, 0.50, 0.05)]),
    "アート2D":             (2.8, 2.8, [(0.50, 0.35, 0.05), (0.76, 0.40, 0.05)]),
    "アニメーション・演出":  (3.6, 2.4, [(0.76, 0.55, 0.045)]),
    "サウンド":             (4.5, 2.6, [(0.76, 0.40, 0.05)]),
    "QA・デバッグ":         (6.0, 1.9, [(0.76, 1.10, 0.045), (0.92, 0.90, 0.03)]),
    "PM・運営":             (1.7, 1.7, []),
}

TYPES = ["新規タイトル開発", "大型アップデート", "移植・リマスター"]

GROUP_SHARE = {
    "新規タイトル開発": {
        "企画・プリプロ": 0.07, "ゲームデザイン": 0.09, "シナリオ": 0.05,
        "プログラム": 0.24, "アート3D": 0.17, "アート2D": 0.08,
        "アニメーション・演出": 0.08, "サウンド": 0.05,
        "QA・デバッグ": 0.11, "PM・運営": 0.06,
    },
    "大型アップデート": {
        "企画・プリプロ": 0.05, "ゲームデザイン": 0.11, "シナリオ": 0.07,
        "プログラム": 0.21, "アート3D": 0.16, "アート2D": 0.09,
        "アニメーション・演出": 0.07, "サウンド": 0.04,
        "QA・デバッグ": 0.14, "PM・運営": 0.06,
    },
    "移植・リマスター": {
        "企画・プリプロ": 0.03, "ゲームデザイン": 0.04, "シナリオ": 0.02,
        "プログラム": 0.38, "アート3D": 0.14, "アート2D": 0.07,
        "アニメーション・演出": 0.05, "サウンド": 0.03,
        "QA・デバッグ": 0.18, "PM・運営": 0.06,
    },
}

CANON_MS = {"プロトタイプ": 0.22, "α版": 0.50, "β版": 0.76, "マスターアップ": 0.92}
MS_SIGMA = {"プロトタイプ": 0.055, "α版": 0.075, "β版": 0.050, "マスターアップ": 0.022}

ORG = {
    "企画・プリプロ":      [("自社", "企画部", "EMP", 1.0)],
    "ゲームデザイン":      [("自社", "ゲームデザイン部", "EMP", 0.8),
                            ("協力会社A", "GDチーム", "PTA", 0.2)],
    "シナリオ":            [("自社", "シナリオ部", "EMP", 0.7),
                            ("協力会社A", "シナリオチーム", "PTA", 0.3)],
    "プログラム":          [("自社", "第1開発部", "EMP", 0.45), ("自社", "第2開発部", "EMP", 0.20),
                            ("協力会社A", "開発支援チーム", "PTA", 0.20),
                            ("協力会社B", "エンジニアリングチーム", "PTB", 0.15)],
    "アート3D":            [("自社", "アート部", "EMP", 0.4),
                            ("協力会社B", "3DCGチーム", "PTB", 0.6)],
    "アート2D":            [("自社", "アート部", "EMP", 0.5),
                            ("協力会社C", "2Dアートチーム", "PTC", 0.5)],
    "アニメーション・演出": [("自社", "アート部", "EMP", 0.5),
                            ("協力会社B", "アニメチーム", "PTB", 0.5)],
    "サウンド":            [("自社", "サウンド部", "EMP", 0.6),
                            ("協力会社C", "サウンドチーム", "PTC", 0.4)],
    "QA・デバッグ":        [("自社", "QA部", "EMP", 0.3),
                            ("協力会社C", "デバッグチーム", "PTC", 0.7)],
    "PM・運営":            [("自社", "PMO", "EMP", 1.0)],
}
MEMBERS_PER_GROUP = {
    "企画・プリプロ": 10, "ゲームデザイン": 16, "シナリオ": 10, "プログラム": 34,
    "アート3D": 34, "アート2D": 20, "アニメーション・演出": 20, "サウンド": 10,
    "QA・デバッグ": 30, "PM・運営": 8,
}
OUTSOURCE_BOOST = 1.6
TAG_POOL = ["コンソール", "モバイル", "PC", "新規IP", "既存IP",
            "オンライン要素", "外注比率高", "海外展開"]


# ==========================================================================
def canonical_density(group, grid):
    a, b, bumps = SHAPE[group]
    y = beta_pdf(grid, a, b)
    y = y / y.max()
    for c, h, w in bumps:
        y = y + h * np.exp(-0.5 * ((grid - c) / w) ** 2)
    return np.clip(y, 0.0, None)


def sample_ms(rng, names):
    while True:
        pos = {n: float(np.clip(CANON_MS[n] + rng.normal(0, MS_SIGMA[n]), 0.05, 0.985))
               for n in names}
        v = [pos[n] for n in names]
        if all(v[i] + 0.04 < v[i + 1] for i in range(len(v) - 1)):
            return pos


def t_to_date(months, t):
    start = pd.Period(months[0], freq="M").start_time
    end = pd.Period(months[-1], freq="M").end_time
    return (start + dt.timedelta(days=float(t) * (end - start).days)).strftime("%Y-%m-%d")


def member_pool(rng, pid, outsourced):
    pool, seq = {}, 0
    for group, orgs in ORG.items():
        entries = []
        for affil, team, prefix, weight in orgs:
            w = weight * (OUTSOURCE_BOOST if (outsourced and affil != "自社") else 1.0)
            n = max(2, int(round(MEMBERS_PER_GROUP[group] * w)))
            for _ in range(n):
                seq += 1
                entries.append((f"{prefix}-{pid[3:]}{seq:04d}", affil, team, w))
        pool[group] = entries
    return pool


def gen_actuals(rng, pid, ptype, mm, months, tags):
    """1案件の実績を作る。戻り値 (DataFrame, マイルストーン実位置)。"""
    n_month = len(months)
    edges = month_edges_t(months)
    names = ["プロトタイプ", "α版", "β版", "マスターアップ"] if ptype == "新規タイトル開発" \
        else ["α版", "β版", "マスターアップ"]
    ms_pos = sample_ms(rng, names)
    xs = [0.0] + [CANON_MS[n] for n in names] + [1.0]
    ys = [0.0] + [ms_pos[n] for n in names] + [1.0]

    total_hours = mm * HOURS_PER_MM * (1.0 + rng.normal(0, 0.012))
    pool = member_pool(rng, pid, "外注比率高" in tags)

    share = np.array([GROUP_SHARE[ptype][g] for g in GROUPS], dtype=float)
    share = share * (1.0 + rng.normal(0, 0.06, size=share.size))
    share = share / share.sum()

    n_grid = 4000
    s_mid = (np.arange(n_grid) + 0.5) / n_grid
    u_mid = np.interp(s_mid, xs, ys)
    bin_idx = np.clip(np.searchsorted(edges, u_mid, side="right") - 1, 0, n_month - 1)

    rec = []
    for gi, group in enumerate(GROUPS):
        dens = canonical_density(group, s_mid)
        dens = dens / dens.sum()
        monthly = np.bincount(bin_idx, weights=dens, minlength=n_month)
        monthly = np.clip(monthly * (1.0 + rng.normal(0, 0.05, size=n_month)), 0, None)
        monthly = monthly / monthly.sum() * total_hours * share[gi]

        entries = pool[group]
        mids = [e[0] for e in entries]
        w = np.array([e[3] for e in entries], dtype=float)
        w = w / w.sum()
        meta = {e[0]: (e[1], e[2]) for e in entries}

        for mi, month in enumerate(months):
            gh = monthly[mi]
            if gh < 1.0:
                continue
            for phase, ratio in PHASES[group]:
                ph = gh * ratio * (1.0 + rng.normal(0, 0.08))
                if ph < 4.0:
                    continue
                n_person = int(np.clip(round(ph / 55.0), 1, len(mids)))
                picked = rng.choice(mids, size=n_person, replace=False, p=w)
                frac = rng.dirichlet(np.full(n_person, 4.0))
                for mid, f in zip(picked, frac):
                    h = round(float(ph * f), 1)
                    if h < 1.0:
                        continue
                    affil, team = meta[mid]
                    rec.append((pid, phase, mid, affil, team, month, h))

    return pd.DataFrame(rec, columns=["案件ID", "行程", "メンバー", "所属",
                                      "チーム", "月", "時間"]), ms_pos


# ==========================================================================
def build_projects(rng, n_done, n_wip):
    """案件マスタを作る。戻り値 (行のリスト, 進行中案件の進捗率)。"""
    rows, progress = [], {}
    year = 2015
    for i in range(n_done):
        ptype = TYPES[i % 3]
        dur = int(rng.integers(12, 37))
        mm = int(np.clip(rng.lognormal(np.log(800), 0.75), 200, 3000))
        start = pd.Period(f"{year + i // 3}-{(i * 5) % 12 + 1:02d}", freq="M")
        end = start + dur - 1
        tags = ";".join(rng.choice(TAG_POOL, size=int(rng.integers(2, 5)), replace=False))
        rows.append((f"PJ-D{i+1:02d}", f"完了案件 {i+1:02d}", ptype, mm,
                     str(start), str(end), tags, "完了"))

    for i in range(n_wip):
        ptype = TYPES[i % 3]
        dur = int(rng.integers(12, 37))
        mm = int(np.clip(rng.lognormal(np.log(800), 0.75), 200, 3000))
        start = pd.Period(f"{2024 + i // 3}-{(i * 4) % 12 + 1:02d}", freq="M")
        # 終了は計画値。実際にはずれる(遅延が多いので上振れ寄り)
        plan_dur = max(6, int(round(dur * (1.0 + rng.normal(0.05, 0.12)))))
        end = start + plan_dur - 1
        tags = ";".join(rng.choice(TAG_POOL, size=int(rng.integers(2, 5)), replace=False))
        rows.append((f"PJ-W{i+1:02d}", f"進行中案件 {i+1:02d}", ptype, mm,
                     str(start), str(end), tags, "進行中"))
        # α版通過直後 〜 マスターアップ直前まで、進捗はばらつかせる
        progress[f"PJ-W{i+1:02d}"] = float(rng.uniform(0.55, 0.88))

    rows.append(("PJ-T01", "予測対象案件", "新規タイトル開発", 1600,
                 "2027-01", "2028-12", "コンソール;新規IP;オンライン要素", "予測対象"))
    return rows, progress


def build_master(rows, ms_rows, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "projects"
    write_sheet(ws, pd.DataFrame(rows, columns=[
        "案件ID", "名称", "種別", "契約人月", "開始", "終了", "タグ", "ステータス"]),
        widths={"案件ID": 12, "名称": 20, "種別": 20, "契約人月": 10,
                "開始": 10, "終了": 10, "タグ": 40, "ステータス": 12},
        input_cols={"契約人月", "開始", "終了", "タグ", "ステータス"},
        note="【凡例】ステータス: 完了 / 進行中 / 予測対象。"
             "進行中案件は実績が途中までしか無く、終了は計画値である点に注意。")

    ws = wb.create_sheet("estimates")
    write_sheet(ws, pd.DataFrame(columns=["案件ID", "行程グループ", "見積人月"]),
                widths={"案件ID": 14, "行程グループ": 26, "見積人月": 12},
                input_cols={"行程グループ", "見積人月"},
                note="【凡例】着手前の案件の要素別見積もり。書いた行程グループだけが予測対象になる。")

    ws = wb.create_sheet("milestones")
    write_sheet(ws, pd.DataFrame(ms_rows, columns=["案件ID", "マイルストーン名", "日付"]),
                widths={"案件ID": 12, "マイルストーン名": 20, "日付": 14},
                input_cols={"マイルストーン名", "日付"},
                note="【凡例】進行中案件の未到達マイルストーンは計画日。")

    ws = wb.create_sheet("phase_map")
    pm = [(p, p, g, MAJOR[g]) for g, ps in PHASES.items() for p, _ in ps]
    write_sheet(ws, pd.DataFrame(pm, columns=["実績行程名", "標準行程名", "行程グループ", "大分類"]),
                widths={"実績行程名": 24, "標準行程名": 24, "行程グループ": 24, "大分類": 12},
                input_cols={"標準行程名", "行程グループ", "大分類"},
                note="【凡例】行程 20 個 / 行程グループ 10 個。")

    ws = wb.create_sheet("settings")
    write_sheet(ws, pd.DataFrame([
        ("人月換算係数", 160, "1人月あたりの時間"),
        ("集約軸", "行程グループ", "phase_map のどの列で学習するか"),
        ("位置合わせ", "ON", "マイルストーンによる位置合わせ"),
        ("背骨マイルストーン", "自動", "位置合わせに使うマイルストーン"),
        ("背骨最小カバー率", 0.6, "この割合以上の案件が持つものを背骨に採用"),
        ("マイルストーン最小件数", 3, "この件数未満は参考値として警告"),
        ("行程グループ最小件数", 3, "この件数未満のカーブは参考値として警告"),
        ("カーブ解像度", 100, "正準時間軸の分割数"),
        ("カーブ復元", "月内均等", "月内均等 / 単調補間"),
        ("k", 3, "種別重み。第1版では未使用"),
        ("タグ重み係数", 0.5, "タグ類似度。第1版では未使用"),
    ], columns=["パラメータ", "値", "説明"]),
        widths={"パラメータ": 24, "値": 16, "説明": 60}, input_cols={"値"})
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data_large"))
    ap.add_argument("--done", type=int, default=30)
    ap.add_argument("--wip", type=int, default=5)
    ap.add_argument("--seed", type=int, default=7)
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)
    rng = np.random.default_rng(a.seed)

    rows, progress = build_projects(rng, a.done, a.wip)
    all_df, ms_rows, summary = [], [], []

    for pid, name, ptype, mm, start, end, tags, status in rows:
        months = month_range(start, end)
        if status == "予測対象":
            for nm in ("α版", "β版"):
                ms_rows.append((pid, nm, t_to_date(months, CANON_MS[nm])))
            summary.append((pid, ptype, mm, len(months), 0, "実績なし"))
            continue

        df, ms_pos = gen_actuals(rng, pid, ptype, mm, months, tags)

        if status == "進行中":
            # 実績を進捗率のところで打ち切る。以降は「まだ記録されていない」
            cut = max(1, int(round(len(months) * progress[pid])))
            keep = set(months[:cut])
            df = df[df["月"].isin(keep)]
            for nm, t in ms_pos.items():
                if nm == "マスターアップ" and t > progress[pid]:
                    continue          # 未到達かつ計画も入れない
                ms_rows.append((pid, nm, t_to_date(months, t)))
            summary.append((pid, ptype, mm, len(months), len(df),
                            f"{df['時間'].sum()/HOURS_PER_MM:.0f}人月 "
                            f"(進捗{progress[pid]:.0%}・{cut}/{len(months)}ヶ月)"))
        else:
            for nm, t in ms_pos.items():
                ms_rows.append((pid, nm, t_to_date(months, t)))
            summary.append((pid, ptype, mm, len(months), len(df),
                            f"{df['時間'].sum()/HOURS_PER_MM:.0f}人月"))
        all_df.append(df)

    actuals = pd.concat(all_df, ignore_index=True)
    actuals = actuals.sort_values(["案件ID", "月", "行程", "メンバー"]).reset_index(drop=True)
    csv = os.path.join(a.out, "actuals.csv")
    actuals.to_csv(csv, index=False, encoding="utf-8-sig")
    build_master(rows, ms_rows, os.path.join(a.out, "master.xlsx"))

    print(f"実績CSV : {csv}  ({len(actuals):,} 行, {os.path.getsize(csv)/1e6:.1f} MB)")
    print(f"マスタ   : {os.path.join(a.out, 'master.xlsx')}")
    print(f"案件     : 完了 {a.done} / 進行中 {a.wip} / 予測対象 1")
    print()
    print(f"{'案件ID':<9}{'種別':<18}{'契約':>6}{'月数':>5}{'行数':>9}  実績")
    for s in summary:
        print(f"{s[0]:<9}{s[1]:<18}{s[2]:>6}{s[3]:>5}{s[4]:>9}  {s[5]}")


if __name__ == "__main__":
    main()
