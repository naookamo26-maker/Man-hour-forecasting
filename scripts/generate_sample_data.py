"""
架空のゲーム会社プロジェクトのサンプルデータを生成する。

出力:
  data/master.xlsx  ... projects / milestones / phase_map / settings
  data/actuals.csv  ... 案件ID / 行程 / メンバー / 所属 / チーム / 月 / 時間

生成の考え方
------------
実データの代わりに、設計書 5章の学習ロジックが「拾えるはず」の構造を
意図的に埋め込んだデータを作る。具体的には:

1. 行程グループごとに正準時間軸(canonical axis)上の工数密度を定義する
   (企画は前寄り、QAは後ろ寄り 等)
2. α版・β版の直前に山(bump)を置く
3. 案件ごとにマイルストーン位置をばらつかせ、正準軸 -> 実時間軸 の
   区分線形ワープでその案件の密度に変換する

=> 経過期間比のまま平均すると 2 の山はぼやけて消える。
   マイルストーンで位置合わせしてから平均すると山が残る。
   すなわち設計書 5章 Step3 の効果が leave-one-out で数値として現れる。
"""

from __future__ import annotations

import datetime as dt
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA_DIR = os.path.join(ROOT, "data")

RNG = np.random.default_rng(20240501)

HOURS_PER_MM = 160  # 人月換算係数(settings と一致させる)

# ----------------------------------------------------------------------------
# 1. 案件マスタ
# ----------------------------------------------------------------------------
# (案件ID, 名称, 種別, 契約人月, 開始年月, 終了年月, タグ, ステータス)
PROJECTS = [
    ("PJ-2019-A", "蒼穹のレガリア",              "新規タイトル開発", 2400, "2019-04", "2022-03", "コンソール;新規IP;オンライン要素", "完了"),
    ("PJ-2020-B", "ブレイブフロンティア外伝",     "大型アップデート", 620,  "2020-01", "2021-02", "モバイル;既存IP",                  "完了"),
    ("PJ-2020-C", "クロノスギア",                "新規タイトル開発", 1800, "2020-07", "2022-12", "コンソール;新規IP;外注比率高",      "完了"),
    ("PJ-2021-D", "蒼穹のレガリア HD Edition",   "移植・リマスター", 520,  "2021-04", "2022-06", "コンソール;既存IP",                "完了"),
    ("PJ-2021-E", "シャドウヴェイル",            "新規タイトル開発", 2950, "2021-01", "2023-12", "コンソール;新規IP;オンライン要素;外注比率高", "完了"),
    ("PJ-2022-F", "ネオンハーツ Season2",        "大型アップデート", 880,  "2022-04", "2023-09", "モバイル;既存IP;オンライン要素",    "完了"),
    ("PJ-2022-G", "ギルドマスターズ",            "新規タイトル開発", 1450, "2022-02", "2024-01", "モバイル;新規IP;オンライン要素",    "完了"),
    ("PJ-2023-H", "クロノスギア:蒼き遺産",       "大型アップデート", 1100, "2023-01", "2024-08", "コンソール;既存IP",                "完了"),
    ("PJ-2023-I", "ブレイブフロンティア Remaster", "移植・リマスター", 700, "2023-05", "2024-10", "コンソール;既存IP;外注比率高",      "完了"),
    ("PJ-2024-J", "オーヴァーロード・サーガ",     "新規タイトル開発", 2200, "2024-01", "2026-06", "コンソール;新規IP;オンライン要素",  "完了"),
    # 予測対象(実績なし)
    ("PJ-2026-K", "蒼穹のレガリア II",           "新規タイトル開発", 1900, "2026-07", "2028-08", "コンソール;既存IP;オンライン要素",  "予測対象"),
]

# ----------------------------------------------------------------------------
# 2. 行程グループと行程
# ----------------------------------------------------------------------------
GROUPS = [
    "企画・プリプロ",
    "ゲームデザイン・シナリオ",
    "プログラム",
    "アート",
    "サウンド",
    "QA・デバッグ",
    "PM・運営",
]

MAJOR = {
    "企画・プリプロ": "上流",
    "ゲームデザイン・シナリオ": "上流",
    "プログラム": "制作",
    "アート": "制作",
    "サウンド": "制作",
    "QA・デバッグ": "検証",
    "PM・運営": "管理",
}

# 行程グループ -> [(実績行程名, グループ内での相対比率)]
PHASES = {
    "企画・プリプロ": [
        ("企画立案", 0.30), ("市場調査", 0.15),
        ("プロトタイプ制作", 0.40), ("企画書作成", 0.15),
    ],
    "ゲームデザイン・シナリオ": [
        ("レベルデザイン", 0.28), ("バトルデザイン", 0.22),
        ("シナリオ執筆", 0.20), ("UI/UX設計", 0.15), ("ゲームバランス調整", 0.15),
    ],
    "プログラム": [
        ("クライアント実装", 0.34), ("サーバ実装", 0.16), ("エンジン開発", 0.14),
        ("ツール開発", 0.11), ("最適化", 0.15), ("移植対応", 0.10),
    ],
    "アート": [
        ("キャラクターモデリング", 0.24), ("背景モデリング", 0.22),
        ("アニメーション", 0.18), ("エフェクト", 0.12),
        ("2Dイラスト", 0.12), ("UIアート", 0.12),
    ],
    "サウンド": [
        ("BGM制作", 0.45), ("SE制作", 0.35), ("ボイス収録", 0.20),
    ],
    "QA・デバッグ": [
        ("デバッグ", 0.62), ("QA計画・進行", 0.23), ("ローカライズQA", 0.15),
    ],
    "PM・運営": [
        ("プロジェクト管理", 0.65), ("進行管理", 0.35),
    ],
}

# 種別ごとの行程グループ工数比率(合計 1.0)
GROUP_SHARE = {
    "新規タイトル開発": {
        "企画・プリプロ": 0.08, "ゲームデザイン・シナリオ": 0.12, "プログラム": 0.27,
        "アート": 0.30, "サウンド": 0.05, "QA・デバッグ": 0.12, "PM・運営": 0.06,
    },
    "大型アップデート": {
        "企画・プリプロ": 0.06, "ゲームデザイン・シナリオ": 0.14, "プログラム": 0.24,
        "アート": 0.32, "サウンド": 0.04, "QA・デバッグ": 0.14, "PM・運営": 0.06,
    },
    "移植・リマスター": {
        "企画・プリプロ": 0.04, "ゲームデザイン・シナリオ": 0.05, "プログラム": 0.40,
        "アート": 0.25, "サウンド": 0.03, "QA・デバッグ": 0.17, "PM・運営": 0.06,
    },
}

# 正準時間軸上の形状: (beta_a, beta_b, [(山の中心, 高さ, 幅), ...])
SHAPE = {
    "企画・プリプロ":           (1.4, 6.0,  []),
    "ゲームデザイン・シナリオ": (2.0, 3.4,  [(0.50, 0.55, 0.05)]),
    "プログラム":               (2.6, 2.4,  [(0.50, 0.80, 0.045), (0.76, 0.45, 0.04)]),
    "アート":                   (3.2, 2.6,  [(0.76, 0.50, 0.05)]),
    "サウンド":                 (4.5, 2.6,  [(0.76, 0.40, 0.05)]),
    "QA・デバッグ":             (6.0, 1.9,  [(0.76, 1.10, 0.045), (0.92, 0.90, 0.03)]),
    "PM・運営":                 (1.7, 1.7,  []),
}

# ----------------------------------------------------------------------------
# 2-2. 消化ペースの型
# ----------------------------------------------------------------------------
# 実データの工数カーブは「中央に大きな山が1つ」ばかりではない。
# 前半が緩やかで後半に一気に消化する案件、早めに片付いて終盤が落ち着く案件、
# 山が2つに割れる案件が実際に存在する。
#
# ここでは「工数の塊をどこへ置くか」を案件ごとに移す形で表現する。
# pace(s) は工数の位置 s を s' に移す単調増加の写像で、pace(0)=0 / pace(1)=1。
# 質量そのものは動かさないので、総工数は変わらず形だけが前後に寄る。
# マイルストーンの位置は動かさないため、
# 「マイルストーンでは説明できない形のばらつき」がデータに残る。
# これが無いと位置合わせの効果を過大評価してしまう。
PACE = {
    "標準":     lambda s: s,
    "後半集中": lambda s: s ** 0.72,      # 工数が後ろへ寄る。前半緩やか、後半に一気
    "前倒し":   lambda s: s ** 1.45,      # 工数が前へ寄る。早めに山、終盤は落ち着く
    "二山":     lambda s: s - 0.11 * np.sin(2 * np.pi * s),   # 中央から離れ、山が2つに割れる
}
# 出現比率。標準が半分、残りを3つで分ける。
PACE_WEIGHTS = {"標準": 0.46, "後半集中": 0.22, "前倒し": 0.22, "二山": 0.10}


def pace_names(n: int, rng) -> list[str]:
    """案件ごとの消化ペースを、比率どおりに割り当てて返す(順序はシャッフル)。"""
    names, weights = list(PACE_WEIGHTS), np.array(list(PACE_WEIGHTS.values()))
    counts = np.maximum(1, np.round(weights / weights.sum() * n).astype(int))
    out = [nm for nm, c in zip(names, counts) for _ in range(c)][:n]
    while len(out) < n:
        out.append("標準")
    rng.shuffle(out)
    return out


# ----------------------------------------------------------------------------
# 3. マイルストーン
# ----------------------------------------------------------------------------
# 正準位置(全案件で共通の「背骨」)
CANON_MS = {"プロトタイプ": 0.22, "α版": 0.50, "β版": 0.76, "マスターアップ": 0.92}

# 種別ごとに存在するマイルストーン
MS_BY_TYPE = {
    "新規タイトル開発": ["プロトタイプ", "α版", "β版", "マスターアップ"],
    "大型アップデート": ["α版", "β版", "マスターアップ"],
    "移植・リマスター": ["α版", "β版", "マスターアップ"],
}

# 案件ごとの実位置ばらつき(標準偏差)
MS_SIGMA = {"プロトタイプ": 0.055, "α版": 0.075, "β版": 0.050, "マスターアップ": 0.022}

# ----------------------------------------------------------------------------
# 4. 所属・チーム
# ----------------------------------------------------------------------------
# 行程グループ -> [(所属, チーム, メンバーID接頭辞, 割当ウェイト)]
ORG = {
    "企画・プリプロ": [
        ("自社", "企画部", "EMP", 1.0),
    ],
    "ゲームデザイン・シナリオ": [
        ("自社", "ゲームデザイン部", "EMP", 0.8),
        ("協力会社A", "GDチーム", "PTA", 0.2),
    ],
    "プログラム": [
        ("自社", "第1開発部", "EMP", 0.45),
        ("自社", "第2開発部", "EMP", 0.2),
        ("協力会社A", "開発支援チーム", "PTA", 0.2),
        ("協力会社B", "エンジニアリングチーム", "PTB", 0.15),
    ],
    "アート": [
        ("自社", "アート部", "EMP", 0.4),
        ("協力会社B", "3DCGチーム", "PTB", 0.35),
        ("協力会社C", "2Dアートチーム", "PTC", 0.25),
    ],
    "サウンド": [
        ("自社", "サウンド部", "EMP", 0.6),
        ("協力会社C", "サウンドチーム", "PTC", 0.4),
    ],
    "QA・デバッグ": [
        ("自社", "QA部", "EMP", 0.3),
        ("協力会社C", "デバッグチーム", "PTC", 0.7),
    ],
    "PM・運営": [
        ("自社", "PMO", "EMP", 1.0),
    ],
}

# 外注比率高タグが付いた案件は協力会社側のウェイトを持ち上げる
OUTSOURCE_BOOST = 1.6


# ============================================================================
# ユーティリティ
# ============================================================================
def month_range(start: str, end: str) -> list[str]:
    """'2021-04' 〜 '2022-06' を月文字列のリストに展開する。"""
    s = pd.Period(start, freq="M")
    e = pd.Period(end, freq="M")
    return [str(p) for p in pd.period_range(s, e, freq="M")]


def month_edges_t(months: list[str]) -> np.ndarray:
    """各月の境界を経過期間比 t(0〜1)で返す。日数ベース。長さ = len(months)+1。"""
    starts = [pd.Period(m, freq="M").start_time for m in months]
    end = pd.Period(months[-1], freq="M").end_time
    days = [(d - starts[0]).days for d in starts] + [(end - starts[0]).days]
    days = np.asarray(days, dtype=float)
    return days / days[-1]


def beta_pdf(x: np.ndarray, a: float, b: float) -> np.ndarray:
    from math import lgamma
    logc = lgamma(a + b) - lgamma(a) - lgamma(b)
    x = np.clip(x, 1e-9, 1 - 1e-9)
    return np.exp(logc + (a - 1) * np.log(x) + (b - 1) * np.log(1 - x))


def canonical_density(group: str, grid: np.ndarray) -> np.ndarray:
    a, b, bumps = SHAPE[group]
    y = beta_pdf(grid, a, b)
    y = y / y.max()
    for center, height, width in bumps:
        y = y + height * np.exp(-0.5 * ((grid - center) / width) ** 2)
    return np.clip(y, 0.0, None)


def sample_milestone_positions(ptype: str) -> dict[str, float]:
    """案件固有のマイルストーン実位置(経過期間比)をサンプリングする。"""
    names = MS_BY_TYPE[ptype]
    while True:
        pos = {}
        for name in names:
            v = CANON_MS[name] + RNG.normal(0, MS_SIGMA[name])
            pos[name] = float(np.clip(v, 0.05, 0.985))
        vals = [pos[n] for n in names]
        if all(vals[i] + 0.04 < vals[i + 1] for i in range(len(vals) - 1)):
            return pos


def make_warp(names: list[str], actual_pos: dict[str, float]):
    """正準時間軸 s -> 実時間軸 u の区分線形ワープを返す。"""
    xs = [0.0] + [CANON_MS[n] for n in names] + [1.0]
    ys = [0.0] + [actual_pos[n] for n in names] + [1.0]
    return lambda s: np.interp(s, xs, ys)


# ============================================================================
# 実績生成
# ============================================================================
def build_member_pool(pid: str, outsourced: bool) -> dict[str, list[tuple]]:
    """行程グループ -> [(メンバーID, 所属, チーム, 選択ウェイト), ...]"""
    pool: dict[str, list[tuple]] = {}
    seq = 0
    for group, orgs in ORG.items():
        entries = []
        for affil, team, prefix, weight in orgs:
            w = weight * (OUTSOURCE_BOOST if (outsourced and affil != "自社") else 1.0)
            n_member = {"企画・プリプロ": 8, "ゲームデザイン・シナリオ": 14,
                        "プログラム": 30, "アート": 40, "サウンド": 8,
                        "QA・デバッグ": 26, "PM・運営": 6}[group]
            n = max(2, int(round(n_member * w)))
            for _ in range(n):
                seq += 1
                mid = f"{prefix}-{pid[-1]}{seq:04d}"
                entries.append((mid, affil, team, w))
        pool[group] = entries
    return pool


def generate_project_actuals(row, pace: str = "標準") -> pd.DataFrame:
    pid, name, ptype, mm, start, end, tags, status = row
    months = month_range(start, end)
    n_month = len(months)
    edges = month_edges_t(months)

    ms_names = MS_BY_TYPE[ptype]
    ms_pos = sample_milestone_positions(ptype)
    warp = make_warp(ms_names, ms_pos)

    total_hours = mm * HOURS_PER_MM * (1.0 + RNG.normal(0, 0.012))
    outsourced = "外注比率高" in tags
    pool = build_member_pool(pid, outsourced)

    # 行程グループ比率(種別ベース + 案件ごとの小さなゆらぎ)
    share = np.array([GROUP_SHARE[ptype][g] for g in GROUPS], dtype=float)
    share = share * (1.0 + RNG.normal(0, 0.06, size=share.size))
    share = share / share.sum()

    # 正準グリッド上で密度を作り、ワープして月次に落とす
    n_grid = 4000
    s_mid = (np.arange(n_grid) + 0.5) / n_grid
    # 消化ペース: 各グリッド点が担う工数はそのままに、置く位置だけを移す。
    # 単調増加で両端を固定しているので、総工数も月数も変わらない。
    s_placed = np.clip(PACE[pace](s_mid), 0.0, 1.0)
    u_mid = warp(s_placed)
    bin_idx = np.clip(np.searchsorted(edges, u_mid, side="right") - 1, 0, n_month - 1)

    records = []
    for gi, group in enumerate(GROUPS):
        dens = canonical_density(group, s_mid)
        dens = dens / dens.sum()
        monthly = np.bincount(bin_idx, weights=dens, minlength=n_month)
        monthly = monthly * (1.0 + RNG.normal(0, 0.05, size=n_month))
        monthly = np.clip(monthly, 0, None)
        monthly = monthly / monthly.sum() * total_hours * share[gi]

        entries = pool[group]
        mids = [e[0] for e in entries]
        weights = np.array([e[3] for e in entries], dtype=float)
        weights = weights / weights.sum()
        meta = {e[0]: (e[1], e[2]) for e in entries}

        for mi, month in enumerate(months):
            gh = monthly[mi]
            if gh < 1.0:
                continue
            for phase, ratio in PHASES[group]:
                ph = gh * ratio * (1.0 + RNG.normal(0, 0.08))
                if ph < 4.0:
                    continue
                # 何人で分担するか(1人あたり月 40〜160h 程度に収まるように)
                n_person = int(np.clip(round(ph / 110.0), 1, len(mids)))
                picked = RNG.choice(mids, size=n_person, replace=False, p=weights)
                w = RNG.dirichlet(np.full(n_person, 4.0))
                for mid, frac in zip(picked, w):
                    h = round(float(ph * frac), 1)
                    if h < 1.0:
                        continue
                    affil, team = meta[mid]
                    records.append((pid, phase, mid, affil, team, month, h))

    return pd.DataFrame(
        records,
        columns=["案件ID", "行程", "メンバー", "所属", "チーム", "月", "時間"],
    ), ms_pos, months, edges


# ============================================================================
# Excel マスタ出力
# ============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def write_sheet(ws, df: pd.DataFrame, widths: dict[str, int] | None = None,
                input_cols: set[str] | None = None, note: str | None = None):
    input_cols = input_cols or set()
    for ci, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, (_, r) in enumerate(df.iterrows(), start=2):
        for ci, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = float(v)
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = INPUT_FONT if col in input_cols else BODY_FONT
    widths = widths or {}
    for ci, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, max(12, len(str(col)) * 2 + 2))
    ws.freeze_panes = "A2"
    if note:
        r = len(df) + 3
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(name="Arial", size=9, italic=True)
        c.fill = NOTE_FILL


def wide_projects_frame(projects_rows, ms_rows) -> pd.DataFrame:
    """案件行 + マイルストーン列 の横持ち表を作る。

    マイルストーン名は見出し行に1回だけ現れるので、表記の揺れが起きない。
    空欄は「未記入」を表し、どの案件に何が入っているかがシート上で一目で分かる。
    同じマイルストーンが複数回ある案件は、1つのセルにカンマ区切りで並べる
    (「2020-04-18, 2020-08-22」)。列は増やさない。
    """
    base = ["案件ID", "名称", "種別", "契約人月", "開始", "終了", "タグ", "ステータス"]
    df = pd.DataFrame(projects_rows, columns=base)
    ms = pd.DataFrame(ms_rows, columns=["案件ID", "マイルストーン名", "日付"])
    if ms.empty:
        return df

    # 列の並びは、案件をまたいだ平均の位置(経過期間比)順にする。
    span = {}
    for _, r in df.iterrows():
        st = pd.Period(str(r["開始"]), freq="M").start_time
        en = pd.Period(str(r["終了"]), freq="M").end_time
        span[str(r["案件ID"])] = (st, max((en - st).days, 1))
    ms["日付"] = pd.to_datetime(ms["日付"])
    ms["位置"] = [((d - span[p][0]).days / span[p][1]) if p in span else 0.5
                  for p, d in zip(ms["案件ID"].astype(str), ms["日付"])]
    order = ms.groupby("マイルストーン名")["位置"].mean().sort_values().index.tolist()

    cells: dict[tuple[str, str], list] = {}
    for (pid, nm), sub in ms.groupby(["案件ID", "マイルストーン名"]):
        cells[(str(pid), nm)] = sorted(sub["日付"].tolist())

    for nm in order:
        df[nm] = [", ".join(d.strftime("%Y-%m-%d") for d in cells.get((pid, nm), [])) or None
                  for pid in df["案件ID"].astype(str)]
    return df


def build_master(projects_rows, ms_rows):
    wb = Workbook()

    # --- projects ---
    ws = wb.active
    ws.title = "projects"
    dfp = wide_projects_frame(projects_rows, ms_rows)
    base = {"案件ID", "名称", "種別", "契約人月", "開始", "終了", "タグ", "ステータス"}
    write_sheet(ws, dfp,
                widths={"案件ID": 14, "名称": 30, "種別": 20, "契約人月": 12,
                        "開始": 10, "終了": 10, "タグ": 40, "ステータス": 12},
                input_cols=({"契約人月", "開始", "終了", "タグ", "ステータス"}
                            | {c for c in dfp.columns if c not in base}),
                note="【凡例】青字セルが手入力項目。ステータス=予測対象 の案件が予測の対象になる"
                     "(実績CSVに行が無くてよい)。"
                     "ステータス列より右はマイルストーン列で、見出しがマイルストーン名、"
                     "セルが日付(YYYY-MM-DD)。空欄=未記入で、記入が無くても動作する(設計書 3-2)。"
                     "同じマイルストーンを一度で通過できず複数回訪れた場合は、"
                     "同じセルにカンマ区切りで並べる(例: 2020-04-18, 2020-08-22)。列は増やさない。"
                     "最終回を工程の境目、初回を「α版(初回)」という別のマイルストーンとして扱う。")

    # --- estimates ---
    # 着手前の案件は実績が1行も無く、分かっているのは要素別の見積もりだけになる。
    # その入力口として空のシートを用意しておく(1行も無ければ契約人月から配分する)。
    ws = wb.create_sheet("estimates")
    write_sheet(ws, pd.DataFrame(columns=["案件ID", "行程グループ", "見積人月"]),
                widths={"案件ID": 14, "行程グループ": 26, "見積人月": 12},
                input_cols={"案件ID", "行程グループ", "見積人月"},
                note="【凡例】着手前の案件は実績が無く、要素別の見積もりだけが分かっている。"
                     "ここに 行程グループ ごとの見積人月を入れると、"
                     "その値がそのままグループ別の総量になる(設計書 6 Step1)。"
                     "行を書かなかった行程グループは『この案件では行わない業務』として"
                     "予測から除外される。1行も無ければ 契約人月 × 学習した工数比率 で配分する。")

    # --- phase_map ---
    ws = wb.create_sheet("phase_map")
    rows = []
    for group, phases in PHASES.items():
        for phase, _ in phases:
            rows.append((phase, phase, group, MAJOR[group]))
    dfm2 = pd.DataFrame(rows, columns=["実績行程名", "標準行程名", "行程グループ", "大分類"])
    write_sheet(ws, dfm2,
                widths={"実績行程名": 24, "標準行程名": 24, "行程グループ": 24, "大分類": 12},
                input_cols={"標準行程名", "行程グループ", "大分類"},
                note="【凡例】青字セルが手入力項目。集約軸の列は増やしてよい(設計書 4-2)。"
                     "settings の 集約軸 をこのシートの列名に変えるだけで学習粒度を切り替えられる。"
                     "表記ゆれのある実績行程名は行を追加して同じ行程グループに寄せる。")

    # --- settings ---
    ws = wb.create_sheet("settings")
    dfs = pd.DataFrame([
        ("人月換算係数", 160, "1人月あたりの時間。出力に必ず明記される(設計書 4-3)"),
        ("集約軸", "行程グループ", "phase_map のどの列で学習するか。大分類 に変えると粗い粒度になる"),
        ("位置合わせ", "ON", "ON/OFF。マイルストーンによる landmark registration(設計書 5 Step3)"),
        ("背骨マイルストーン", "自動", "位置合わせに使うマイルストーン名を ; 区切りで指定。自動=全案件共通のものを使う"),
        ("背骨最小カバー率", 0.6, "この割合以上の案件が持つマイルストーンを背骨に採用する。1.0=全案件必須"),
        ("位置合わせ強度", 1.0, "0〜1。実位置を正準位置へ引き戻す割合。下げると時間軸の伸縮が緩み、工数の跳ねが減る"),
        ("伸縮率上限", 0, "時間軸の伸縮率の上限(倍)。例 1.5。0=無制限。跳ねの高さに直接上限をかける"),
        ("マイルストーン最小件数", 3, "この件数未満の統計は参考値として警告する(設計書 11)"),
        ("カーブ解像度", 100, "正準時間軸の分割数。学習カーブのビン数"),
        ("k", 3, "種別重み w=n/(n+k)。第1版では未使用。実装順序 3 で使用する"),
        ("タグ重み係数", 0.5, "タグ類似度の効き。第1版では未使用。実装順序 4 で使用する"),
    ], columns=["パラメータ", "値", "説明"])
    write_sheet(ws, dfs,
                widths={"パラメータ": 24, "値": 16, "説明": 70},
                input_cols={"値"},
                note="【凡例】青字セルが手入力項目。k とタグ重み係数は設計書 9 の実装順序 3・4 で使用予定。"
                     "第1版(素朴版+位置合わせ)では読み込むだけで計算に使わない。")

    return wb


# ============================================================================
def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    all_actuals = []
    ms_rows = []
    summary = []
    project_rows = []

    done = [r for r in PROJECTS if r[7] != "予測対象"]
    paces = dict(zip([r[0] for r in done], pace_names(len(done), RNG)))

    for row in PROJECTS:
        pid, name, ptype, mm, start, end, tags, status = row
        pace = paces.get(pid, "標準")
        # 消化ペースは名称に括弧書きで残す。どの案件がどの形かを見て確かめられるようにする。
        name = name if status == "予測対象" or pace == "標準" else f"{name}({pace})"
        row = (pid, name, ptype, mm, start, end, tags, status)
        project_rows.append(row)
        months = month_range(start, end)

        if status == "予測対象":
            # 実績なし。α版だけ人が指定済み、という想定で1点だけ入れる
            ms_pos = {"α版": 0.48}
            for msname, t in ms_pos.items():
                ms_rows.append((pid, msname, t_to_date(months, t)))
            summary.append((pid, name, ptype, mm, len(months), 0, "予測対象"))
            continue

        df, ms_pos, months, edges = generate_project_actuals(row, pace)
        all_actuals.append(df)
        for msname in MS_BY_TYPE[ptype]:
            ms_rows.append((pid, msname, t_to_date(months, ms_pos[msname])))
        summary.append((pid, name, ptype, mm, len(months), len(df),
                        f"{df['時間'].sum() / HOURS_PER_MM:.0f}人月 / 消化ペース {pace}"))

    actuals = pd.concat(all_actuals, ignore_index=True)
    actuals = actuals.sort_values(["案件ID", "月", "行程", "メンバー"]).reset_index(drop=True)
    out_csv = os.path.join(DATA_DIR, "actuals.csv")
    actuals.to_csv(out_csv, index=False, encoding="utf-8-sig")

    wb = build_master(project_rows, ms_rows)
    out_xlsx = os.path.join(DATA_DIR, "master.xlsx")
    wb.save(out_xlsx)

    print(f"実績CSV : {out_csv}  ({len(actuals):,} 行, {os.path.getsize(out_csv)/1e6:.1f} MB)")
    print(f"マスタ   : {out_xlsx}")
    print()
    print(f"{'案件ID':<12}{'名称':<28}{'種別':<20}{'契約':>7}{'月数':>6}{'行数':>9}  実績")
    for s in summary:
        print(f"{s[0]:<12}{s[1]:<28}{s[2]:<20}{s[3]:>7}{s[4]:>6}{s[5]:>9}  {s[6]}")


def t_to_date(months: list[str], t: float) -> str:
    """経過期間比 t を実際の日付(YYYY-MM-DD)に変換する。"""
    start = pd.Period(months[0], freq="M").start_time
    end = pd.Period(months[-1], freq="M").end_time
    d = start + dt.timedelta(days=float(t) * (end - start).days)
    return d.strftime("%Y-%m-%d")


if __name__ == "__main__":
    main()
