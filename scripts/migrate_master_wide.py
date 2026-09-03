"""master.xlsx の milestones シートを projects シートの右側へ統合する。

これまではマイルストーンを別シートに 案件ID / マイルストーン名 / 日付 の3列で
書いていたが、手入力ではシート間を往復することになり、
マイルストーン名の表記も案件ごとにぶれる。

新しい形式では、projects シートの案件情報の右にマイルストーンを列として並べ、
見出し行をマイルストーン名にする。

    案件ID | 名称 | 種別 | 契約人月 | 開始 | 終了 | タグ | ステータス | α版 | β版 | マスターアップ
    PJ-01  | ...  | ...  |     800 | ...  | ...  | ... | 完了       | 2021-05-10 |  | 2022-03-01

  - マイルストーン名は見出し行で1回決まるので、表記の揺れが起きない
  - 空欄 = 未記入。どの案件に何が入っているかがひと目で分かる
  - 案件1行を書けば入力が終わる

同じマイルストーンを複数回通過した場合(通らずにやり直した場合)は、
見出しを α版 / α版#2 / α版#3 と増やす。

    python scripts/migrate_master_wide.py                 # data/master.xlsx を変換
    python scripts/migrate_master_wide.py --master path/to/master.xlsx
    python scripts/migrate_master_wide.py --drop-old      # 旧シートを残さない

変換前のファイルは <名前>.bak.xlsx として残す。
"""

from __future__ import annotations

import argparse
import os
import shutil
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_loader import PROJECT_BASE_COLS

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF7E6")
HEADER_FONT = Font(bold=True)


def _data_row_count(ws) -> int:
    """見出しの次の行から、最初の空行までのデータ行数を返す。

    シート末尾の凡例(注記)を巻き込まないため。読み込み層の _read_table と同じ規則。
    """
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            break
        n += 1
    return n


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="milestones シートを projects シートへ統合する")
    ap.add_argument("--master", default=os.path.join(ROOT, "data", "master.xlsx"))
    ap.add_argument("--drop-old", action="store_true",
                    help="旧 milestones シートを残さずに削除する(既定は milestones_旧 にリネーム)")
    a = ap.parse_args(argv)

    if not os.path.exists(a.master):
        print(f"[中断] マスタが見つかりません: {a.master}")
        return 1

    wb = load_workbook(a.master)
    if "milestones" not in wb.sheetnames:
        print("milestones シートがありません。既に統合済みか、マイルストーン未入力です。")
        return 0
    if "projects" not in wb.sheetnames:
        print("[中断] projects シートがありません。")
        return 1

    ms = pd.read_excel(a.master, sheet_name="milestones")
    ms = ms.dropna(subset=["案件ID", "マイルストーン名", "日付"])
    ms["案件ID"] = ms["案件ID"].astype(str).str.strip()
    ms["マイルストーン名"] = ms["マイルストーン名"].astype(str).str.strip()
    ms["日付"] = pd.to_datetime(ms["日付"], errors="coerce")
    ms = ms.dropna(subset=["日付"])
    if ms.empty:
        print("milestones シートに有効な行がありません。")
        return 0

    ws = wb["projects"]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    n_rows = _data_row_count(ws)
    pids = [str(ws.cell(row=r, column=1).value).strip() for r in range(2, 2 + n_rows)]

    already = [h for h in header if h and h not in PROJECT_BASE_COLS]
    if already:
        print(f"[中断] projects シートに既にマイルストーン列があります: {', '.join(already)}")
        print("       二重に足さないため、ここで止めます。")
        return 1

    # 列の並びは「案件をまたいだ平均の位置」順にする。
    # 案件ごとに期間が違うので日付そのままでは並ばない。開始からの経過期間比で測る。
    projects = pd.read_excel(a.master, sheet_name="projects").head(n_rows)
    span = {}
    for _, r in projects.iterrows():
        try:
            st = pd.Period(str(r["開始"]), freq="M").start_time
            en = pd.Period(str(r["終了"]), freq="M").end_time
            span[str(r["案件ID"]).strip()] = (st, max((en - st).days, 1))
        except Exception:
            continue

    def _pos(row):
        s = span.get(row["案件ID"])
        return (row["日付"] - s[0]).days / s[1] if s else 0.5

    ms["位置"] = ms.apply(_pos, axis=1)
    order = ms.groupby("マイルストーン名")["位置"].mean().sort_values().index.tolist()

    # 同じ案件に同じマイルストーンが複数回あれば、#2 / #3 と列を増やす。
    max_rep = {nm: int(ms[ms["マイルストーン名"] == nm]
                       .groupby("案件ID").size().max()) for nm in order}
    columns: list[tuple[str, str, int]] = []      # (列見出し, 名前, 何回目)
    for nm in order:
        for i in range(1, max_rep[nm] + 1):
            columns.append((nm if i == 1 else f"{nm}#{i}", nm, i))

    table = {(pid, nm): [] for pid in pids for nm in order}
    for (pid, nm), sub in ms.groupby(["案件ID", "マイルストーン名"]):
        if (pid, nm) in table:
            table[(pid, nm)] = sorted(sub["日付"].tolist())

    unknown = sorted(set(ms["案件ID"]) - set(pids))
    if unknown:
        print(f"[警告] projects に無い案件のマイルストーンは移行できません: "
              f"{', '.join(unknown[:8])}" + (" ほか" if len(unknown) > 8 else ""))

    base_n = len(header)
    for j, (label, nm, i) in enumerate(columns):
        col = base_n + j + 1
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(12, len(label) + 4)
        for k, pid in enumerate(pids):
            dates = table.get((pid, nm), [])
            c = ws.cell(row=2 + k, column=col)
            c.fill = INPUT_FILL
            c.number_format = "yyyy-mm-dd"
            if len(dates) >= i:
                c.value = dates[i - 1].to_pydatetime()

    if a.drop_old:
        del wb["milestones"]
    else:
        wb["milestones"].title = "milestones_旧"

    backup = os.path.splitext(a.master)[0] + ".bak.xlsx"
    shutil.copy2(a.master, backup)
    wb.save(a.master)

    n_filled = sum(1 for pid in pids for nm in order if table.get((pid, nm)))
    print(f"変換しました: {a.master}")
    print(f"  バックアップ : {backup}")
    print(f"  追加した列   : {len(columns)} 列  ({', '.join(c[0] for c in columns)})")
    print(f"  記入済みセル : {sum(len(table.get((pid, nm), [])) for pid in pids for nm in order)} 件"
          f" / {len(pids)} 案件中 {len({pid for pid in pids if any(table.get((pid, nm)) for nm in order)})} 案件に記入あり")
    repeats = [(pid, nm, len(v)) for (pid, nm), v in table.items() if len(v) > 1]
    if repeats:
        print(f"  複数回のマイルストーン: {len(repeats)} 件")
        for pid, nm, n in sorted(repeats, key=lambda x: -x[2])[:10]:
            print(f"    {pid} {nm}: {n} 回")
    if not a.drop_old:
        print("  旧 milestones シートは milestones_旧 に改名して残しました。"
              "確認できたら削除してください。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
