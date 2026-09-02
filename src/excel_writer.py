"""
Excel 出力(設計書 7章)

シート構成
  予測   月 × 行程グループ の表(主役) + 積み上げグラフ
  カーブ  学習した平均カーブと各案件のカーブの重ね描き
  検証   leave-one-out の結果
  条件   使用パラメータ、依拠した案件、換算係数

出力方針
  - 合計は数式(SUM)で書く。値を疑ったとき、その場で検算できる状態を保つ。
  - 人月表示は「時間 ÷ 換算係数」の数式にし、係数セルを参照させる。
    係数を書き換えれば表全体が追従する。数字だけが独り歩きしない。
  - 換算係数は 予測シートと条件シートの両方に明記する(設計書 4-3)。
"""

from __future__ import annotations

import datetime as dt

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F3864")
H2_FONT = Font(name=FONT, size=11, bold=True, color="1F3864")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="7F7F7F")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
def _title(ws, row, text, font=TITLE_FONT):
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    return row + 1


def _note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    return row + 1


def _table(ws, row, df: pd.DataFrame, *, index_header: str | None = None,
           num_fmt: dict | None = None, col_width: dict | None = None,
           start_col: int = 1) -> tuple[int, int, int]:
    """DataFrame を書き出す。戻り値 = (次の行, ヘッダ行, データ最終行)。"""
    num_fmt = num_fmt or {}
    col_width = col_width or {}
    head_row = row
    cols = list(df.columns)
    if not cols:
        # 該当データが無い場合。落とさず、その旨だけ残して次へ進む。
        c = ws.cell(row=head_row, column=start_col, value="(該当データなし)")
        c.font = NOTE_FONT
        return head_row + 2, head_row, head_row

    ci0 = start_col
    if index_header is not None:
        c = ws.cell(row=head_row, column=ci0, value=index_header)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BOX
        ci0 += 1

    for j, col in enumerate(cols):
        c = ws.cell(row=head_row, column=ci0 + j, value=str(col))
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BOX

    for i, (idx, r) in enumerate(df.iterrows()):
        rr = head_row + 1 + i
        cc = start_col
        if index_header is not None:
            c = ws.cell(row=rr, column=cc, value=idx)
            c.font, c.border = BODY_FONT, BOX
            cc += 1
        for j, col in enumerate(cols):
            v = r[col]
            if isinstance(v, np.integer):
                v = int(v)
            elif isinstance(v, np.floating):
                v = None if np.isnan(v) else float(v)
            c = ws.cell(row=rr, column=cc + j, value=v)
            c.font, c.border = BODY_FONT, BOX
            if col in num_fmt:
                c.number_format = num_fmt[col]

    # 列幅
    if index_header is not None:
        ws.column_dimensions[get_column_letter(start_col)].width = \
            col_width.get(index_header, max(10, len(index_header) * 2))
    for j, col in enumerate(cols):
        letter = get_column_letter(ci0 + j)
        ws.column_dimensions[letter].width = col_width.get(col, max(11, len(str(col)) * 1.9))

    last = head_row + len(df)
    return last + 2, head_row, last


# ===========================================================================
# 予測シート
# ===========================================================================
def _sheet_forecast(wb, fc, model, ds, cond_ref: str, actual_table=None):
    ws = wb.create_sheet("予測")
    # 予測した行程グループだけを扱う。この案件が行わない業務は列ごと落ちている。
    fgroups = fc.groups
    r = 1
    r = _title(ws, r, f"工数予測  {fc.name}  ({fc.pid})")
    ws.cell(row=r, column=1, value="生成日時").font = BODY_FONT
    ws.cell(row=r, column=2, value=dt.datetime.now().strftime("%Y-%m-%d %H:%M")).font = BODY_FONT
    r += 1

    prow = ds.project(fc.pid)
    for label, val, fmt in [
        ("種別", prow["種別"], None),
        ("タグ", prow.get("タグ", ""), None),
        ("期間", f"{fc.months[0]} 〜 {fc.months[-1]}  ({len(fc.months)}ヶ月)", None),
        ("契約人月", float(prow["契約人月"]), "#,##0"),
        ("人月換算係数", fc.hours_per_mm, "#,##0"),
        ("総工数(時間)", fc.total_hours, "#,##0"),
    ]:
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BODY_FONT
        if fmt:
            c.number_format = fmt
        if label == "人月換算係数":
            c.fill = KEY_FILL
            ws.cell(row=r, column=3,
                    value="← 人月表示はこのセルを参照。定義が違う場合はここを書き換える").font = NOTE_FONT
            coef_cell = f"$B${r}"
        r += 1
    r += 1

    # --- 主表(時間) ---
    r = _title(ws, r, "■ 月 × 行程グループ 工数(時間)", H2_FONT)
    # 表示用に丸めると合計が総工数から数値分ずれる。
    # 「表の合計 = 契約人月 × 換算係数」は手で検算される数字なので、丸めた後に差を吸収する。
    tbl = fc.table.round(1)
    resid = round(fc.total_hours - tbl.to_numpy().sum(), 1)
    if abs(resid) >= 0.1:
        biggest = tbl.sum(axis=1).idxmax()
        gmax = tbl.loc[biggest].idxmax()
        tbl.loc[biggest, gmax] = round(tbl.loc[biggest, gmax] + resid, 1)
    tbl.insert(0, "経過月", range(1, len(tbl) + 1))
    r_next, head, last = _table(
        ws, r, tbl, index_header="月",
        num_fmt={g: "#,##0" for g in fgroups} | {"経過月": "0"},
        col_width={"月": 10, "経過月": 8} | {g: 15 for g in fgroups})

    n_g = len(fgroups)
    g_first = 3            # C列 = 最初の行程グループ
    g_last = g_first + n_g - 1
    tot_col = g_last + 1

    # 行合計・列合計を数式で
    c = ws.cell(row=head, column=tot_col, value="合計")
    c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
    c.alignment = Alignment(horizontal="center")
    for rr in range(head + 1, last + 1):
        c = ws.cell(row=rr, column=tot_col,
                    value=f"=SUM({get_column_letter(g_first)}{rr}:{get_column_letter(g_last)}{rr})")
        c.font, c.border, c.number_format = BODY_FONT, BOX, "#,##0"

    sum_row = last + 1
    c = ws.cell(row=sum_row, column=1, value="合計")
    c.font, c.border = Font(name=FONT, size=10, bold=True), BOX
    for cc in range(g_first, tot_col + 1):
        L = get_column_letter(cc)
        c = ws.cell(row=sum_row, column=cc, value=f"=SUM({L}{head+1}:{L}{last})")
        c.font, c.border, c.number_format = Font(name=FONT, size=10, bold=True), BOX, "#,##0"
        c.fill = GOOD_FILL
    ratio_row = sum_row + 1
    c = ws.cell(row=ratio_row, column=1, value="構成比")
    c.font, c.border = NOTE_FONT, BOX
    for cc in range(g_first, g_last + 1):
        L = get_column_letter(cc)
        c = ws.cell(row=ratio_row, column=cc,
                    value=f"={L}{sum_row}/${get_column_letter(tot_col)}${sum_row}")
        c.font, c.border, c.number_format = NOTE_FONT, BOX, "0.0%"

    hours_head, hours_first, hours_last = head, head + 1, last
    r = ratio_row + 3

    # --- 人月表示(数式で時間表から換算) ---
    r = _title(ws, r, "■ 月 × 行程グループ 工数(人月)", H2_FONT)
    _note(ws, r, f"上の時間表を換算係数({int(fc.hours_per_mm)}時間/人月)で割った値。"
                 "係数セルを書き換えると自動で追従する。")
    r += 1
    mm_head = r
    c = ws.cell(row=mm_head, column=1, value="月")
    c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
    for j, g in enumerate(fgroups):
        c = ws.cell(row=mm_head, column=2 + j, value=g)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    c = ws.cell(row=mm_head, column=2 + n_g, value="合計")
    c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX

    for i in range(len(fc.months)):
        rr = mm_head + 1 + i
        src = hours_first + i
        c = ws.cell(row=rr, column=1, value=f"=A{src}")
        c.font, c.border = BODY_FONT, BOX
        for j in range(n_g):
            L = get_column_letter(g_first + j)
            c = ws.cell(row=rr, column=2 + j, value=f"={L}{src}/{coef_cell}")
            c.font, c.border, c.number_format = BODY_FONT, BOX, "#,##0.0"
        c = ws.cell(row=rr, column=2 + n_g,
                    value=f"=SUM(B{rr}:{get_column_letter(1+n_g)}{rr})")
        c.font, c.border, c.number_format = BODY_FONT, BOX, "#,##0.0"
    mm_last = mm_head + len(fc.months)
    c = ws.cell(row=mm_last + 1, column=1, value="合計")
    c.font, c.border = Font(name=FONT, size=10, bold=True), BOX
    for j in range(n_g + 1):
        L = get_column_letter(2 + j)
        c = ws.cell(row=mm_last + 1, column=2 + j, value=f"=SUM({L}{mm_head+1}:{L}{mm_last})")
        c.font, c.border, c.number_format = Font(name=FONT, size=10, bold=True), BOX, "#,##0.0"
        c.fill = GOOD_FILL
    r = mm_last + 3

    # --- 実績との比較 ---
    # 予測対象が進行中なら、同じ月軸・同じ集計軸で実績を並べる。
    # 予測だけを見ても、当たっているのか外れているのかは判断できない。
    r = _title(ws, r, "■ 実績(時間)と 予測との比較", H2_FONT)
    has_actual = actual_table is not None and not actual_table.empty \
        and float(actual_table.to_numpy().sum()) > 0
    if not has_actual:
        r = _note(ws, r, "この案件の実績はまだ1行もありません。"
                         "着手後に実績が入ると、ここに 月 × 行程グループ の実績と"
                         "予測との差分が並びます。")
        r += 1
    else:
        act = actual_table.reindex(index=fc.months, columns=fgroups).fillna(0.0).round(1)
        r = _note(ws, r, "実績は予測と同じ集計(phase_map に載る行程・契約期間内)で作っている。"
                         "実績が入っている月までが比較対象で、それ以降の実績は 0 と表示される。")
        r += 1
        # 上の予測表と列位置を揃える(経過月の分だけ行程グループが右へずれる)。
        # 揃えておかないと合計式の参照範囲が1列ずれ、先頭グループが抜ける。
        act.insert(0, "経過月", range(1, len(act) + 1))
        r_next, ahead, alast = _table(
            ws, r, act, index_header="月",
            num_fmt={g: "#,##0" for g in fgroups} | {"経過月": "0"},
            col_width={"月": 10, "経過月": 8} | {g: 15 for g in fgroups})
        c = ws.cell(row=ahead, column=tot_col, value="合計")
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
        for rr in range(ahead + 1, alast + 1):
            c = ws.cell(row=rr, column=tot_col,
                        value=f"=SUM({get_column_letter(g_first)}{rr}:"
                              f"{get_column_letter(g_last)}{rr})")
            c.font, c.border, c.number_format = BODY_FONT, BOX, "#,##0"
        a_sum = alast + 1
        c = ws.cell(row=a_sum, column=1, value="合計")
        c.font, c.border = Font(name=FONT, size=10, bold=True), BOX
        for cc in range(g_first, tot_col + 1):
            L = get_column_letter(cc)
            c = ws.cell(row=a_sum, column=cc, value=f"=SUM({L}{ahead+1}:{L}{alast})")
            c.font, c.border, c.number_format = Font(name=FONT, size=10, bold=True), BOX, "#,##0"
            c.fill = GOOD_FILL
        r = a_sum + 3

        # --- 月次の 予測 vs 実績(差分は数式にして、どちらの表を直しても追従させる) ---
        r = _title(ws, r, "■ 月次 予測 vs 実績(合計・時間)", H2_FONT)
        d_head = r
        TL = get_column_letter(tot_col)
        for j, lab in enumerate(["月", "予測", "実績", "差分(予測-実績)", "消化率(実績/予測)"]):
            c = ws.cell(row=d_head, column=1 + j, value=lab)
            c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(1 + j)].width = max(12, len(lab) * 2)
        for i in range(len(fc.months)):
            rr = d_head + 1 + i
            fsrc, asrc = hours_first + i, ahead + 1 + i
            vals = [f"=A{fsrc}", f"={TL}{fsrc}", f"={TL}{asrc}",
                    f"={TL}{fsrc}-{TL}{asrc}",
                    f'=IF({TL}{fsrc}=0,"",{TL}{asrc}/{TL}{fsrc})']
            for j, v in enumerate(vals):
                c = ws.cell(row=rr, column=1 + j, value=v)
                c.font, c.border = BODY_FONT, BOX
                c.number_format = "0.0%" if j == 4 else ("#,##0" if j else "General")
        d_last = d_head + len(fc.months)
        c = ws.cell(row=d_last + 1, column=1, value="合計")
        c.font, c.border = Font(name=FONT, size=10, bold=True), BOX
        for j in range(1, 4):
            L = get_column_letter(1 + j)
            c = ws.cell(row=d_last + 1, column=1 + j,
                        value=f"=SUM({L}{d_head+1}:{L}{d_last})")
            c.font, c.border, c.number_format = Font(name=FONT, size=10, bold=True), BOX, "#,##0"
            c.fill = GOOD_FILL
        c = ws.cell(row=d_last + 1, column=5, value=f"=IF(B{d_last+1}=0,\"\",C{d_last+1}/B{d_last+1})")
        c.font, c.border, c.number_format = Font(name=FONT, size=10, bold=True), BOX, "0.0%"
        c.fill = GOOD_FILL

        cmp_chart = LineChart()
        cmp_chart.title = "月次工数 予測 vs 実績(時間)"
        cmp_chart.y_axis.title, cmp_chart.x_axis.title = "時間", "月"
        cmp_chart.height, cmp_chart.width = 9, min(30, 8 + 0.5 * len(fc.months))
        cmp_chart.add_data(Reference(ws, min_col=2, max_col=3,
                                     min_row=d_head, max_row=d_last), titles_from_data=True)
        cmp_chart.set_categories(Reference(ws, min_col=1, min_row=d_head + 1, max_row=d_last))
        for s in cmp_chart.series:
            s.smooth = False
        ws.add_chart(cmp_chart, f"G{d_head}")
        r = d_last + 3

    # --- マイルストーン ---
    r = _title(ws, r, "■ マイルストーン(予測)", H2_FONT)
    r, _, _ = _table(ws, r, fc.milestones,
                     num_fmt={"位置t": "0.000"},
                     col_width={"マイルストーン名": 20, "予測日付": 14, "位置t": 10, "根拠": 22})
    for n in fc.notes:
        r = _note(ws, r, "・" + n)
    r = _note(ws, r, f"・この予測が依拠した案件は「条件」シートを参照(→ {cond_ref})")

    # --- グラフ ---
    ch = BarChart()
    ch.type, ch.grouping, ch.overlap = "col", "stacked", 100
    ch.title = f"{fc.name} 月次工数(時間・行程グループ積み上げ)"
    ch.y_axis.title = "時間"
    ch.x_axis.title = "月"
    ch.height, ch.width = 10, min(38, 8 + 0.7 * len(fc.months))
    data = Reference(ws, min_col=g_first, max_col=g_last,
                     min_row=hours_head, max_row=hours_last)
    cats = Reference(ws, min_col=1, min_row=hours_first, max_row=hours_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, f"{get_column_letter(tot_col + 2)}{hours_head}")

    ln = LineChart()
    ln.title = "累積消化率"
    ln.y_axis.title = "累積比率"
    ln.height, ln.width = 8, min(30, 8 + 0.5 * len(fc.months))
    cum_col = tot_col + 1
    ws.cell(row=hours_head, column=cum_col, value="累積率").fill = HEAD_FILL
    ws.cell(row=hours_head, column=cum_col).font = HEAD_FONT
    TL = get_column_letter(tot_col)
    for i in range(len(fc.months)):
        rr = hours_first + i
        c = ws.cell(row=rr, column=cum_col,
                    value=f"=SUM(${TL}${hours_first}:{TL}{rr})/${TL}${sum_row}")
        c.font, c.border, c.number_format = BODY_FONT, BOX, "0.0%"
    ln.add_data(Reference(ws, min_col=cum_col, min_row=hours_head, max_row=hours_last),
                titles_from_data=True)
    ln.set_categories(cats)
    ws.add_chart(ln, f"{get_column_letter(tot_col + 2)}{hours_head + 21}")

    ws.freeze_panes = "C" + str(hours_head + 1)
    return ws


# ===========================================================================
# カーブシート
# ===========================================================================
def _sheet_curves(wb, model, model_naive, curves, ds):
    ws = wb.create_sheet("カーブ")
    r = 1
    r = _title(ws, r, "学習カーブ")
    r = _note(ws, r, "横軸は正準時間軸(0=開始, 1=終了)。位置合わせ ON のとき、"
                     "背骨マイルストーンが全案件で同じ位置に来るよう変換した後の軸。")
    r = _note(ws, r, f"背骨マイルストーン: {', '.join(model.backbone) or '(なし)'}"
                     f"   正準位置: "
                     f"{', '.join(f'{k}={v:.3f}' for k, v in model.canonical_anchor.items()) or '-'}")
    r += 1

    n = model.n_bin
    t = (np.arange(n) + 0.5) / n

    # --- 累積消化カーブ ---
    r = _title(ws, r, "■ 累積消化率 F(t):平均カーブ vs 各案件", H2_FONT)
    data = {"平均(位置合わせあり)": np.cumsum(model.total_shape)}
    if model_naive is not None:
        data["平均(位置合わせなし)"] = np.cumsum(model_naive.total_shape)
    for pid in sorted(model.project_shapes):
        data[f"{pid}"] = np.cumsum(model.project_shapes[pid])
    df = pd.DataFrame(data, index=np.round(t, 4))
    df.index.name = "t"
    r_next, head, last = _table(ws, r, df, index_header="t",
                                num_fmt={c: "0.000" for c in df.columns},
                                col_width={"t": 8} | {c: 13 for c in df.columns})

    ln = LineChart()
    ln.title = "累積消化率カーブ(太線=平均)"
    ln.y_axis.title, ln.x_axis.title = "累積比率", "正準時間軸 t"
    ln.height, ln.width = 11, 24
    ln.add_data(Reference(ws, min_col=2, max_col=1 + len(df.columns),
                          min_row=head, max_row=last), titles_from_data=True)
    ln.set_categories(Reference(ws, min_col=1, min_row=head + 1, max_row=last))
    for s in ln.series:
        s.marker.symbol = "none"
        s.smooth = False
    ws.add_chart(ln, f"{get_column_letter(len(df.columns) + 3)}{head}")

    r = r_next + 22

    # --- 案件別の密度(非累積) ---
    # 累積カーブは単調増加なので、どの案件も似たS字に見えて差が読み取りにくい。
    # 山がどこに立つか、いくつあるか、どれだけ鋭いかは非累積で初めて見える。
    r = _title(ws, r, "■ 工数密度(非累積):平均カーブ vs 各案件", H2_FONT)
    r = _note(ws, r, "上の累積カーブと同じデータを、累積せずに density として並べたもの。"
                     "面積が1になるよう正規化しているので、案件規模によらず形だけを比べられる。"
                     "山の位置・鋭さ・数の違いはこちらでしか見えない。")
    r += 1
    dens_p = {"平均(位置合わせあり)": model.total_shape * n}
    if model_naive is not None:
        dens_p["平均(位置合わせなし)"] = model_naive.total_shape * n
    for pid in sorted(model.project_shapes):
        dens_p[pid] = model.project_shapes[pid] * n
    dfd = pd.DataFrame(dens_p, index=np.round(t, 4))
    dfd.index.name = "t"
    r_next, headd, lastd = _table(ws, r, dfd, index_header="t",
                                  num_fmt={c: "0.000" for c in dfd.columns},
                                  col_width={"t": 8} | {c: 13 for c in dfd.columns})
    lnd = LineChart()
    lnd.title = "工数密度カーブ(非累積・太線=平均)"
    lnd.y_axis.title, lnd.x_axis.title = "密度", "正準時間軸 t"
    lnd.height, lnd.width = 11, 24
    lnd.add_data(Reference(ws, min_col=2, max_col=1 + len(dfd.columns),
                           min_row=headd, max_row=lastd), titles_from_data=True)
    lnd.set_categories(Reference(ws, min_col=1, min_row=headd + 1, max_row=lastd))
    for s in lnd.series:
        s.marker.symbol = "none"
        s.smooth = False
    ws.add_chart(lnd, f"{get_column_letter(len(dfd.columns) + 3)}{headd}")

    r = r_next + 22

    # --- 行程グループ別の密度 ---
    r = _title(ws, r, "■ 行程グループ別 平均カーブ(密度・面積1)", H2_FONT)
    r = _note(ws, r, "各行程グループが正準時間軸のどこで消化されるか。"
                     "山が立つ位置がマイルストーン直前に来ていれば位置合わせが効いている。")
    dens = pd.DataFrame({g: model.shape[g] * n for g in model.groups},
                        index=np.round(t, 4))
    dens.index.name = "t"
    r_next, head2, last2 = _table(ws, r, dens, index_header="t",
                                  num_fmt={g: "0.000" for g in model.groups},
                                  col_width={"t": 8} | {g: 15 for g in model.groups})
    ln2 = LineChart()
    ln2.title = "行程グループ別 工数密度(平均カーブ)"
    ln2.y_axis.title, ln2.x_axis.title = "密度", "正準時間軸 t"
    ln2.height, ln2.width = 11, 24
    ln2.add_data(Reference(ws, min_col=2, max_col=1 + len(model.groups),
                           min_row=head2, max_row=last2), titles_from_data=True)
    ln2.set_categories(Reference(ws, min_col=1, min_row=head2 + 1, max_row=last2))
    for s in ln2.series:
        s.marker.symbol = "none"
    ws.add_chart(ln2, f"{get_column_letter(len(model.groups) + 3)}{head2}")

    r = r_next + 22
    r = _title(ws, r, "■ 案件ごとのワープ(実位置 → 正準位置)", H2_FONT)
    rows = []
    for pid, c in sorted(curves.items()):
        row = {"案件ID": pid, "名称": c.name, "種別": c.ptype, "期間(月)": len(c.months)}
        for nm in model.ms_stats["マイルストーン名"]:
            row[nm] = round(c.ms_t[nm], 3) if nm in c.ms_t else None
        rows.append(row)
    wdf = pd.DataFrame(rows)
    anchor = {"案件ID": "(正準位置)", "名称": "", "種別": "", "期間(月)": None}
    for nm in model.ms_stats["マイルストーン名"]:
        anchor[nm] = round(model.canonical_anchor[nm], 3) if nm in model.canonical_anchor else None
    wdf = pd.concat([wdf, pd.DataFrame([anchor])], ignore_index=True)
    _table(ws, r, wdf, num_fmt={nm: "0.000" for nm in model.ms_stats["マイルストーン名"]},
           col_width={"案件ID": 13, "名称": 26, "種別": 18})
    return ws


# ===========================================================================
# 検証シート
# ===========================================================================
def _sheet_validation(wb, detail, monthly, summary):
    ws = wb.create_sheet("検証")
    r = 1
    r = _title(ws, r, "leave-one-out 検証")
    r = _note(ws, r, "1件を除いた残りで学習し、除いた1件を予測して実績と比較する。"
                     "工数の量ではなく形を評価するため、予測値は実績合計に正規化してから比較している。")
    r = _note(ws, r, "WAPE = Σ|予測-実績| / Σ実績。0 が完全一致。")
    r += 1

    r = _title(ws, r, "■ モード別サマリ", H2_FONT)
    fmt = {c: "0.000" for c in summary.columns if "WAPE" in c or "乖離" in c}
    fmt["ピーク月ズレ_絶対値平均"] = "0.0"
    fmt["素朴版比_改善率"] = "0.0%"
    r_next, head, last = _table(ws, r, summary, num_fmt=fmt,
                                col_width={"モード": 22, "案件数": 8})
    ws.cell(row=head, column=1).alignment = Alignment(horizontal="center")
    r = r_next
    r = _note(ws, r, "素朴版 vs 位置合わせ(MS学習値):平均カーブを鋭くしただけで当たるようになるかを測る。")
    r = _note(ws, r, "位置合わせ(MS学習値) vs 位置合わせ(MS指定):マイルストーンを手入力する価値を測る。")
    r += 1

    r = _title(ws, r, "■ 案件別 誤差", H2_FONT)
    piv = detail.pivot(index="案件ID", columns="モード", values="月次誤差WAPE")
    piv = piv.reindex(columns=[m for m in summary["モード"] if m in piv.columns])
    r_next, head2, last2 = _table(ws, r, piv, index_header="案件ID",
                                  num_fmt={c: "0.000" for c in piv.columns},
                                  col_width={"案件ID": 13} | {c: 20 for c in piv.columns})
    ch = BarChart()
    ch.type, ch.grouping = "col", "clustered"
    ch.title = "案件別 月次誤差WAPE(低いほど良い)"
    ch.y_axis.title = "WAPE"
    ch.height, ch.width = 10, 24
    ch.add_data(Reference(ws, min_col=2, max_col=1 + len(piv.columns),
                          min_row=head2, max_row=last2), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=head2 + 1, max_row=last2))
    ws.add_chart(ch, f"{get_column_letter(len(piv.columns) + 3)}{head2}")
    r = r_next + 20

    r = _title(ws, r, "■ 案件別 詳細指標", H2_FONT)
    d = detail.copy()
    r_next, _, _ = _table(ws, r, d,
                          num_fmt={"月次誤差WAPE": "0.000", "月×グループ誤差WAPE": "0.000",
                                   "累積カーブ最大乖離": "0.000", "ピーク高さ誤差": "0.0%",
                                   "契約人月": "#,##0"},
                          col_width={"案件ID": 13, "名称": 26, "種別": 18, "モード": 22})
    r = r_next

    r = _title(ws, r, "■ 月次 予測 vs 実績(全案件・全モード)", H2_FONT)
    r = _note(ws, r, "ピボットテーブルで案件・モードを切り替えて確認できる形にしてある。")
    _table(ws, r, monthly,
           num_fmt={"予測(時間)": "#,##0", "実績(時間)": "#,##0", "差分(時間)": "#,##0",
                    "予測累積率": "0.0%", "実績累積率": "0.0%"},
           col_width={"案件ID": 13, "モード": 22, "月": 10})
    return ws


# ===========================================================================
# 条件シート
# ===========================================================================
def _sheet_conditions(wb, model, ds, fc, params: dict, warnings: list[str], curves):
    ws = wb.create_sheet("条件")
    r = 1
    r = _title(ws, r, "実行条件")
    r = _note(ws, r, "パラメータを変えて何度も回す段階では、"
                     "「この結果はどの設定で出したか」を追えないと結果の比較ができなくなる。")
    r += 1

    r = _title(ws, r, "■ パラメータ", H2_FONT)
    pdf = pd.DataFrame([{"項目": k, "値": v} for k, v in params.items()])
    r, _, _ = _table(ws, r, pdf, col_width={"項目": 26, "値": 60})

    r = _title(ws, r, "■ 人月換算", H2_FONT)
    c = ws.cell(row=r, column=1, value="換算係数")
    c.font = BODY_FONT
    c = ws.cell(row=r, column=2, value=model.hours_per_mm)
    c.font, c.fill, c.number_format = BODY_FONT, KEY_FILL, "#,##0"
    ws.cell(row=r, column=3, value="時間 / 人月").font = BODY_FONT
    r += 1
    r = _note(ws, r, "人月の定義は組織によって揺れる。この係数を明記しないまま数字が独り歩きすると"
                     "議論が噛み合わなくなる(設計書 4-3)。")
    r += 1

    r = _title(ws, r, "■ 学習に使用した案件と重み", H2_FONT)
    r = _note(ws, r, "第1版(素朴版+位置合わせ)では全案件の重みを 1.0 で固定している。"
                     "種別・タグによる重み付けは実装順序 3・4 で導入する。")
    r, _, _ = _table(ws, r, model.contributors,
                     num_fmt={"重み": "0.00", "契約人月": "#,##0", "実績人月": "#,##0"},
                     col_width={"案件ID": 13, "名称": 28, "種別": 18})

    # 上の表の「実績人月」を行程グループへ分解したもの。
    # 学習された工数比率がどの案件のどの業務から来ているかを、ここで追える。
    r = _title(ws, r, "■ 案件別 × 行程グループ 実績工数(人月)", H2_FONT)
    r = _note(ws, r, "予測と同じ集計(phase_map に載る行程・契約期間内)での実績合計。"
                     "0 はその案件がその業務を行っていないことを示す。"
                     "最下行は全案件の合計で、その構成比が学習値そのものではない点に注意"
                     "(学習は案件ごとの構成比を平均するため、規模の大きい案件に引っ張られない)。")
    r += 1
    rows = []
    for pid, c in sorted(curves.items()):
        s = c.monthly.sum(axis=0).reindex(model.groups).fillna(0.0) / model.hours_per_mm
        row = {"案件ID": pid, "名称": c.name, "期間(月)": len(c.months)}
        row.update({g: round(float(s[g]), 1) for g in model.groups})
        row["合計"] = round(float(s.sum()), 1)
        rows.append(row)
    if rows:
        gdf_p = pd.DataFrame(rows)
        total_row = {"案件ID": "合計", "名称": "", "期間(月)": None}
        total_row.update({g: round(float(gdf_p[g].sum()), 1) for g in model.groups})
        total_row["合計"] = round(float(gdf_p["合計"].sum()), 1)
        share_row = {"案件ID": "構成比", "名称": "(全案件の単純合計)", "期間(月)": None}
        tot = total_row["合計"] or 1.0
        share_row.update({g: round(total_row[g] / tot, 4) for g in model.groups})
        share_row["合計"] = 1.0
        gdf_p = pd.concat([gdf_p, pd.DataFrame([total_row, share_row])], ignore_index=True)
        r, ghead, glast = _table(ws, r, gdf_p,
                                 num_fmt={g: "#,##0.0" for g in model.groups} | {"合計": "#,##0.0"},
                                 col_width={"案件ID": 13, "名称": 26, "期間(月)": 9}
                                 | {g: 15 for g in model.groups})
        # 最終行(構成比)だけは比率なので、列単位の書式を上書きして%表示にする。
        for cc in range(4, 4 + len(model.groups) + 1):
            cell = ws.cell(row=glast, column=cc)
            cell.number_format = "0.0%"
            cell.font = NOTE_FONT

    r = _title(ws, r, "■ マイルストーン位置分布(学習値)", H2_FONT)
    r, _, _ = _table(ws, r, model.ms_stats,
                     num_fmt={"平均位置": "0.000", "標準偏差": "0.000"},
                     col_width={"マイルストーン名": 22})

    r = _title(ws, r, "■ 行程グループ別 工数比率(学習値)", H2_FONT)
    r = _note(ws, r, "件数 = そのグループの形状カーブの学習に参加した案件数。"
                     "そのグループの業務が無い案件は形の平均に参加しないため、"
                     "件数が少ないカーブは参考値として扱うこと。"
                     "比率のほうは業務が無い案件も0として平均に含めている。")
    gdf = pd.DataFrame({"行程グループ": model.group_ratio.index,
                        "比率": model.group_ratio.to_numpy(),
                        "件数": [model.group_sample_n.get(g, 0)
                                 for g in model.group_ratio.index]})
    r, _, _ = _table(ws, r, gdf, num_fmt={"比率": "0.0%", "件数": "0"},
                     col_width={"行程グループ": 26, "比率": 12, "件数": 8})

    if warnings:
        r = _title(ws, r, "■ 警告", H2_FONT)
        for w in warnings:
            c = ws.cell(row=r, column=1, value="⚠ " + w)
            c.font = Font(name=FONT, size=10, color="C00000")
            r += 1
    return ws


# ===========================================================================
def _page_setup(ws):
    """関係者にそのまま渡せるよう、印刷しても崩れない設定にしておく。"""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False


def write_workbook(path: str, *, fc, model, model_naive, curves, ds,
                   detail, monthly, summary, params: dict, warnings: list[str],
                   actual_table=None):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_forecast(wb, fc, model, ds, "条件", actual_table=actual_table)
    _sheet_curves(wb, model, model_naive, curves, ds)
    if not summary.empty:
        _sheet_validation(wb, detail, monthly, summary)
    _sheet_conditions(wb, model, ds, fc, params, warnings, curves)
    for ws in wb.worksheets:
        _page_setup(ws)
    wb.save(path)
    return path
